"""Generate, validate and export a synthetic textile operations dataset."""
from pathlib import Path
import json
import numpy as np
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import LineChart, Reference

ROOT = Path(__file__).resolve().parents[1]
DATA = ROOT / "data"
EXCEL = ROOT / "excel"
DATA.mkdir(exist_ok=True); EXCEL.mkdir(exist_ok=True)
rng = np.random.default_rng(42)
n = 1500
dates = pd.to_datetime(rng.choice(pd.date_range("2025-01-01", "2025-12-31"), n))
plants = rng.choice(["Delhi NCR", "Surat", "Ludhiana", "Jaipur"], n, p=[.27,.30,.24,.19])
lines = [f"L{rng.integers(1,4):02d}" for _ in range(n)]
shifts = rng.choice(["Morning","Evening","Night"], n, p=[.38,.35,.27])
fabrics = rng.choice(["Cotton","Denim","Polyester","Viscose","Blend"], n, p=[.31,.20,.19,.14,.16])
target = rng.integers(780, 1220, n)
eff = np.clip(rng.normal(.92, .055, n) - np.where(shifts=="Night",.018,0), .72, 1.05)
actual = np.round(target*eff, 1)
defect_rate = np.clip(rng.normal(.032, .013, n)+np.where(shifts=="Night",.006,0), .006, .095)
defects = np.round(actual*defect_rate,1); good=np.round(actual-defects,1)
issued=np.round(actual/rng.uniform(3.1,3.8,n),1); waste=np.round(issued*np.clip(rng.normal(.039,.012,n),.01,.09),1)
downtime=np.clip(np.round(rng.gamma(2.2,16,n)+np.where(eff<.86,20,0)),0,150).astype(int)
prices=pd.Series(fabrics).map({"Cotton":132,"Denim":175,"Polyester":108,"Viscose":148,"Blend":124}).to_numpy()
on_time=rng.random(n)<np.clip(.94-(.9-eff)*.75-(downtime>70)*.14,.55,.98)
df=pd.DataFrame({
 "run_id":[f"TR{str(i+1).zfill(5)}" for i in range(n)],"production_date":dates,"month":dates.strftime("%b"),
 "plant":plants,"production_line":[f"{p[:2].upper()}-{l}" for p,l in zip(plants,lines)],"shift":shifts,"fabric_type":fabrics,
 "customer_segment":rng.choice(["Domestic Brand","Exporter","Wholesaler","Fashion House"],n),"target_meters":target,
 "actual_meters":actual,"good_meters":good,"defect_meters":defects,"defect_type":rng.choice(["Weaving","Shade Variation","Stain","Yarn Break","Finishing"],n,p=[.27,.20,.16,.22,.15]),
 "material_issued_kg":issued,"waste_kg":waste,"downtime_minutes":downtime,"machine_speed_rpm":rng.integers(620,910,n),
 "operators_count":rng.integers(5,13,n),"energy_kwh":np.round(actual*rng.uniform(.18,.28,n),1),
 "order_value_inr":np.round(actual*prices,2),"delivery_status":np.where(on_time,"On Time","Late"),
 "quality_grade":np.select([defect_rate<.025,defect_rate<.05],["A","B"],default="C")})
df=df.sort_values(["production_date","run_id"]).reset_index(drop=True)
df["production_date"]=df["production_date"].dt.strftime("%Y-%m-%d")
df.to_csv(DATA/"textile_production_data.csv",index=False)

summary={"records":len(df),"production_meters":round(df.actual_meters.sum(),1),"efficiency_pct":round(100*df.actual_meters.sum()/df.target_meters.sum(),2),"defect_rate_pct":round(100*df.defect_meters.sum()/df.actual_meters.sum(),2),"waste_rate_pct":round(100*df.waste_kg.sum()/df.material_issued_kg.sum(),2),"on_time_pct":round(100*(df.delivery_status=="On Time").mean(),2),"order_value_inr":round(df.order_value_inr.sum(),2)}
(DATA/"dashboard_summary.json").write_text(json.dumps(summary,indent=2),encoding="utf-8")

kpis=pd.DataFrame({"KPI":["Production (m)","Efficiency %","Defect Rate %","Waste Rate %","On-Time Delivery %","Order Value (INR)"],"Value":[summary["production_meters"],summary["efficiency_pct"],summary["defect_rate_pct"],summary["waste_rate_pct"],summary["on_time_pct"],summary["order_value_inr"]]})
monthly=df.groupby("month",sort=False).agg(Target_Meters=("target_meters","sum"),Actual_Meters=("actual_meters","sum"),Defect_Meters=("defect_meters","sum"),Downtime_Minutes=("downtime_minutes","sum")).reset_index()
dictionary=pd.read_html((ROOT/"docs/DATA_DICTIONARY.md").read_text().replace("|---|---|",""))[0] if False else pd.DataFrame({"Field":df.columns,"Description":["See docs/DATA_DICTIONARY.md"]*len(df.columns)})
path=EXCEL/"Textile_Operations_Analytics.xlsx"
with pd.ExcelWriter(path,engine="openpyxl") as writer:
 df.to_excel(writer,index=False,sheet_name="Raw Data");kpis.to_excel(writer,index=False,sheet_name="KPI Summary");monthly.to_excel(writer,index=False,sheet_name="Monthly Trend");dictionary.to_excel(writer,index=False,sheet_name="Data Dictionary")
 wb=writer.book
 for ws in wb.worksheets:
  ws.freeze_panes="A2";ws.auto_filter.ref=ws.dimensions
  for cell in ws[1]: cell.font=Font(bold=True,color="FFFFFF");cell.fill=PatternFill("solid",fgColor="0B7057");cell.alignment=Alignment(horizontal="center")
  for col in ws.columns: ws.column_dimensions[col[0].column_letter].width=min(max(len(str(c.value or "")) for c in col)+2,28)
 ws=wb["Monthly Trend"];chart=LineChart();chart.title="Monthly Production vs Target";chart.y_axis.title="Metres";chart.x_axis.title="Month";chart.add_data(Reference(ws,min_col=2,max_col=3,min_row=1,max_row=13),titles_from_data=True);chart.set_categories(Reference(ws,min_col=1,min_row=2,max_row=13));ws.add_chart(chart,"F2")
print(json.dumps(summary,indent=2));print(f"Created {path}")

