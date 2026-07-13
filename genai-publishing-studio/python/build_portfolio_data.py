"""Build synthetic production data, SQLite DB, dashboard JSON and Excel workbook."""
from pathlib import Path
import csv, json, random, sqlite3
from statistics import mean

ROOT=Path(__file__).resolve().parents[1]; DATA=ROOT/'data'; DATA.mkdir(exist_ok=True)
random.seed(42)
modalities=['Image','Video','Voice','Animation','Lecture']
subjects=['EVS','English','Hindi','GK','Science','Mathematics']
rows=[]
for i in range(1,121):
    modality=random.choice(modalities); ai_hours=round(random.uniform(1.2,8.0),1)
    manual=round(ai_hours*random.uniform(1.7,3.5),1); revisions=random.choices([0,1,2,3],[54,28,13,5])[0]
    quality=round(min(99,max(72,random.gauss(90-revisions*1.4,3.8))),1)
    rows.append({'asset_id':f'GA-{i:03}','modality':modality,'subject':random.choice(subjects),'quality_score':quality,'revisions':revisions,'ai_hours':ai_hours,'estimated_manual_hours':manual,'hours_saved':round(manual-ai_hours,1),'approved_first_pass':int(revisions==0)})
fields=list(rows[0]);
with (DATA/'production_log.csv').open('w',newline='',encoding='utf-8') as f:
    w=csv.DictWriter(f,fieldnames=fields);w.writeheader();w.writerows(rows)
db=sqlite3.connect(DATA/'genai_studio.db');db.execute('DROP TABLE IF EXISTS production_assets');
db.execute('CREATE TABLE production_assets(asset_id TEXT PRIMARY KEY, modality TEXT, subject TEXT, quality_score REAL, revisions INTEGER, ai_hours REAL, estimated_manual_hours REAL, hours_saved REAL, approved_first_pass INTEGER)')
db.executemany('INSERT INTO production_assets VALUES (?,?,?,?,?,?,?,?,?)',[tuple(r[k] for k in fields) for r in rows]);db.commit();db.close()
by=[]
for m in modalities:
    s=[r for r in rows if r['modality']==m];by.append({'modality':m,'assets':len(s),'quality':round(mean(r['quality_score'] for r in s),1)})
metrics={'assets':len(rows),'first_pass_approval':round(100*mean(r['approved_first_pass'] for r in rows),1),'avg_quality':round(mean(r['quality_score'] for r in rows),1),'hours_saved':round(sum(r['hours_saved'] for r in rows),1),'by_modality':by}
(DATA/'dashboard_metrics.json').write_text(json.dumps(metrics,indent=2),encoding='utf-8')
try:
    from openpyxl import Workbook
    from openpyxl.chart import BarChart,Reference
    from openpyxl.styles import Font,PatternFill
    out=ROOT/'excel';out.mkdir(exist_ok=True);wb=Workbook();ws=wb.active;ws.title='Production Log';ws.append(fields)
    for r in rows:ws.append([r[k] for k in fields])
    for c in ws[1]:c.font=Font(bold=True,color='FFFFFF');c.fill=PatternFill('solid',fgColor='6747FF')
    ws.freeze_panes='A2';ws.auto_filter.ref=ws.dimensions
    dash=wb.create_sheet('Dashboard');dash.append(['Modality','Assets','Avg Quality'])
    for x in by:dash.append([x['modality'],x['assets'],x['quality']])
    chart=BarChart();chart.title='Quality by Modality';chart.add_data(Reference(dash,min_col=3,min_row=1,max_row=6),titles_from_data=True);chart.set_categories(Reference(dash,min_col=1,min_row=2,max_row=6));dash.add_chart(chart,'E2')
    wb.save(out/'GenAI_Studio_Analytics.xlsx')
except ImportError: print('openpyxl missing: Excel file skipped')
print(json.dumps(metrics,indent=2))
