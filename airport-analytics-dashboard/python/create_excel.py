"""Create a polished Excel workbook from the analysis-ready CSV."""
from pathlib import Path
import pandas as pd
from openpyxl.styles import Font, PatternFill

ROOT=Path(__file__).resolve().parents[1]
df=pd.read_csv(ROOT/"data"/"airport_operations.csv")
df["on_time"]=df.departure_delay_min.le(15)
df["load_factor"]=df.passengers/df.seat_capacity
kpis=pd.DataFrame({"KPI":["Total Flights","Total Passengers","On-Time %","Average Delay","Load Factor %"],"Value":[len(df),df.passengers.sum(),df.on_time.mean(),df.departure_delay_min.mean(),df.load_factor.mean()]})
airline=df.groupby("airline").agg(Flights=("flight_id","count"),Passengers=("passengers","sum"),On_Time=("on_time","mean"),Avg_Delay=("departure_delay_min","mean"),Load_Factor=("load_factor","mean")).reset_index()
terminal=df.groupby("terminal").agg(Flights=("flight_id","count"),Passengers=("passengers","sum"),Security_Wait=("security_wait_min","mean"),Baggage_Delivery=("baggage_delivery_min","mean")).reset_index()
out=ROOT/"excel"/"Airport_Analytics.xlsx"
with pd.ExcelWriter(out,engine="openpyxl") as writer:
    df.to_excel(writer,index=False,sheet_name="Flight Data"); kpis.to_excel(writer,index=False,sheet_name="KPI Summary"); airline.to_excel(writer,index=False,sheet_name="Airline Performance"); terminal.to_excel(writer,index=False,sheet_name="Terminal Performance")
    for ws in writer.book.worksheets:
        ws.freeze_panes="A2"; ws.auto_filter.ref=ws.dimensions
        for cell in ws[1]: cell.font=Font(color="FFFFFF",bold=True); cell.fill=PatternFill("solid",fgColor="123B5D")
        for col in ws.columns: ws.column_dimensions[col[0].column_letter].width=min(28,max(12,max(len(str(c.value or "")) for c in col)+2))
print(f"Created {out}")
